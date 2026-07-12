"""Add Skipped_38 and Existing_in_DB_545 sheets to Combined.xlsx."""
import json
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
FILE = ROOT / "Combined.xlsx"
DB_JSON = ROOT / "scripts" / "_db_books.json"
REQUIRED = ["اسم الكتاب", "المؤلف", "القسم", "الصندوق"]
SHEET_SKIPPED = "Skipped_38"
SHEET_DUPLICATES = "Existing_in_DB_545"


def norm(s):
    return (str(s).strip().lower() if s is not None and str(s).strip() != "" else "")


def get(row, idx, col):
    i = idx.get(col)
    if i is None or i >= len(row):
        return ""
    v = row[i]
    if v is None:
        return ""
    return str(v).strip()


def row_nonempty(row):
    return any(c is not None and str(c).strip() != "" for c in row)


def copy_row_values(row, width):
    values = list(row) if row else []
    if len(values) < width:
        values.extend([None] * (width - len(values)))
    return values[:width]


def autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def main():
    db_books = json.loads(DB_JSON.read_text(encoding="utf-8"))
    db_index = {}
    for b in db_books:
        key = (norm(b.get("name")), norm(b.get("author")), norm(b.get("publisher")))
        db_index.setdefault(key, []).append(b)

    wb = openpyxl.load_workbook(FILE)
    source = wb[wb.sheetnames[0]]
    source_rows = list(source.iter_rows(values_only=True))
    if not source_rows:
        raise SystemExit("Source sheet is empty")

    header = [str(c).strip() if c is not None else "" for c in source_rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    header_width = len(header)

    skipped_rows = []
    duplicate_rows = []
    for excel_row_num, row in enumerate(source_rows[1:], start=2):
        if not row_nonempty(row):
            continue
        miss = [f for f in REQUIRED if not get(row, idx, f)]
        if miss:
            skipped_rows.append((excel_row_num, miss, row))
            continue
        key = (
            norm(get(row, idx, "اسم الكتاب")),
            norm(get(row, idx, "المؤلف")),
            norm(get(row, idx, "دار النشر")),
        )
        if key in db_index:
            duplicate_rows.append((excel_row_num, row))

    for name in (SHEET_SKIPPED, SHEET_DUPLICATES):
        if name in wb.sheetnames:
            del wb[name]

    ws_skip = wb.create_sheet(SHEET_SKIPPED)
    ws_skip.append(header + ["Excel_Row", "Missing_Fields"])
    for row_num, miss, row in skipped_rows:
        ws_skip.append(copy_row_values(row, header_width) + [row_num, ", ".join(miss)])
    autosize_columns(ws_skip)

    ws_dup = wb.create_sheet(SHEET_DUPLICATES)
    ws_dup.append(header + ["Excel_Row"])
    for row_num, row in duplicate_rows:
        ws_dup.append(copy_row_values(row, header_width) + [row_num])
    autosize_columns(ws_dup)

    wb.save(FILE)
    summary = {
        "skipped": len(skipped_rows),
        "duplicates": len(duplicate_rows),
        "sheets": wb.sheetnames,
    }
    Path(ROOT / "scripts" / "_combined_sheets_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
