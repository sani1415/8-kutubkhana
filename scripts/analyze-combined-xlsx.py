"""Analyze Combined.xlsx for Kutubkhana CSV import compatibility."""
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
FILE = ROOT / "Combined.xlsx"
DB_JSON = ROOT / "scripts" / "_db_books.json"
REPORT = ROOT / "scripts" / "_combined_analysis.json"
EXPECTED = [
    "اسم الكتاب", "المؤلف", "القسم", "المحقق", "الأجزاء", "دار النشر",
    "السنة", "النسخ", "الحالة", "الصندوق", "الطاق", "ملاحظات",
]
REQUIRED = ["اسم الكتاب", "المؤلف", "القسم", "الصندوق"]


def norm(s):
    return (str(s).strip().lower() if s is not None and str(s).strip() != "" else "")


def year_valid(v):
    s = str(v).strip() if v is not None else ""
    if not s or s.lower() == "nan":
        return True
    s2 = re.sub(r"[\s\u0660-\u0669\u06F0-\u06F9]", "", s)
    digits_only = re.sub(r"\d", "", s2)
    return len(digits_only) == 0 and len(s) <= 8


def main():
    wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}

    def get(row, col):
        i = idx.get(col)
        if i is None or i >= len(row):
            return ""
        v = row[i]
        if v is None:
            return ""
        return str(v).strip()

    data_rows = rows[1:]
    missing_required = []
    valid_rows = []
    for i, row in enumerate(data_rows, start=2):
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue
        miss = [f for f in REQUIRED if not get(row, f)]
        if miss:
            missing_required.append({
                "row": i,
                "missing": miss,
                "name": get(row, "اسم الكتاب"),
                "author": get(row, "المؤلف"),
            })
        else:
            valid_rows.append((i, row))

    seen = {}
    internal_dups = []
    for i, row in valid_rows:
        key = (norm(get(row, "اسم الكتاب")), norm(get(row, "المؤلف")), norm(get(row, "دار النشر")))
        if key in seen:
            internal_dups.append({"rows": [seen[key], i], "name": get(row, "اسم الكتاب"), "author": get(row, "المؤلف"), "publisher": get(row, "دار النشر")})
        else:
            seen[key] = i

    seen2 = {}
    internal_dups_na = []
    for i, row in valid_rows:
        key = (norm(get(row, "اسم الكتاب")), norm(get(row, "المؤلف")))
        if key in seen2:
            internal_dups_na.append({"rows": [seen2[key], i], "name": get(row, "اسم الكتاب"), "author": get(row, "المؤلف")})
        else:
            seen2[key] = i

    bad_years = []
    for i, row in valid_rows:
        y = get(row, "السنة")
        if y and not year_valid(y):
            bad_years.append({"row": i, "year": y, "name": get(row, "اسم الكتاب")})

    statuses = Counter(get(r, "الحالة") or "(empty)" for _, r in valid_rows)
    cats = Counter(get(r, "القسم") for _, r in valid_rows)
    cabs = Counter(get(r, "الصندوق") for _, r in valid_rows)

    non_numeric_parts = []
    non_numeric_copies = []
    for i, row in valid_rows:
        p = get(row, "الأجزاء")
        if p and not re.fullmatch(r"\d+", p.replace(" ", "")):
            non_numeric_parts.append({"row": i, "value": p, "name": get(row, "اسم الكتاب")})
        c = get(row, "النسخ")
        if c and not re.fullmatch(r"\d+", c.replace(" ", "")):
            non_numeric_copies.append({"row": i, "value": c, "name": get(row, "اسم الكتاب")})

    report = {
        "file": str(FILE),
        "sheets": wb.sheetnames,
        "header": header,
        "expected_columns": EXPECTED,
        "missing_expected_columns": [c for c in EXPECTED if c not in header],
        "extra_columns": [h for h in header if h and h not in EXPECTED],
        "header_first_12_exact_match": header[:12] == EXPECTED,
        "total_data_rows": len(data_rows),
        "non_empty_rows": sum(1 for r in data_rows if any(c is not None and str(c).strip() != "" for c in r)),
        "valid_importable_rows": len(valid_rows),
        "rows_missing_required": len(missing_required),
        "missing_required_samples": missing_required[:25],
        "internal_duplicates_name_author_publisher": len(internal_dups),
        "internal_duplicates_samples": internal_dups[:20],
        "internal_duplicates_name_author_only": len(internal_dups_na),
        "internal_duplicates_na_samples": internal_dups_na[:15],
        "invalid_years": bad_years,
        "status_distribution": dict(statuses),
        "unexpected_status": [s for s in statuses if s not in {"متاح", "معار", "(empty)", ""}],
        "unique_categories": len(cats),
        "top_categories": cats.most_common(12),
        "unique_cabinets": len(cabs),
        "top_cabinets": cabs.most_common(12),
        "non_numeric_parts": non_numeric_parts,
        "non_numeric_copies": non_numeric_copies,
    }

    if DB_JSON.exists():
        db_books = json.loads(DB_JSON.read_text(encoding="utf-8"))
        db_index = {}
        db_index_na = {}
        for b in db_books:
            k = (norm(b.get("name")), norm(b.get("author")), norm(b.get("publisher")))
            db_index.setdefault(k, []).append(b)
            k2 = (norm(b.get("name")), norm(b.get("author")))
            db_index_na.setdefault(k2, []).append(b)

        exact_matches = []
        would_update = []
        new_books = []
        for i, row in valid_rows:
            name = get(row, "اسم الكتاب")
            author = get(row, "المؤلف")
            publisher = get(row, "دار النشر")
            key = (norm(name), norm(author), norm(publisher))
            key_na = (norm(name), norm(author))
            if key in db_index:
                exact_matches.append({"row": i, "name": name, "author": author, "publisher": publisher})
            elif key_na in db_index_na:
                would_update.append({
                    "row": i, "name": name, "author": author, "publisher": publisher,
                    "db_matches": len(db_index_na[key_na]),
                })
            else:
                new_books.append({"row": i, "name": name, "author": author, "publisher": publisher})

        report["database"] = {
            "total_books_in_db": len(db_books),
            "exact_duplicate_name_author_publisher": len(exact_matches),
            "same_name_author_different_publisher": len(would_update),
            "truly_new_books": len(new_books),
            "exact_duplicate_samples": exact_matches[:20],
            "new_book_samples": new_books[:20],
            "name_author_diff_publisher_samples": would_update[:15],
        }

    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    wb.close()
    print(str(REPORT))


if __name__ == "__main__":
    main()
