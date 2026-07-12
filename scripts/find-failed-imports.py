"""Find failed imports using pre-import DB snapshot (3975 books, no id)."""
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
FILE = ROOT / "Combined.xlsx"
DB_JSON = ROOT / "scripts" / "_db_books.json"
REQUIRED = ["اسم الكتاب", "المؤلف", "القسم", "الصندوق"]


def norm(s):
    return (str(s).strip().lower() if s is not None and str(s).strip() != "" else "")


def main():
    db_books = json.loads(DB_JSON.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}

    def get(row, col):
        i = idx.get(col)
        if i is None or i >= len(row) or row[i] is None:
            return ""
        return str(row[i]).strip()

    existing = [
        {
            "id": f"db-{i}",
            "name": b.get("name", ""),
            "author": b.get("author", ""),
            "publisher": b.get("publisher", ""),
        }
        for i, b in enumerate(db_books)
    ]

    failed = []
    tasks_log = []

    for row_num, row in enumerate(rows[1:], start=2):
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue
        if any(not get(row, f) for f in REQUIRED):
            continue

        name = get(row, "اسم الكتاب")
        author = get(row, "المؤلف")
        publisher = get(row, "دار النشر")
        key = (norm(name), norm(author), norm(publisher))

        existing_book = next(
            (
                b
                for b in existing
                if norm(b.get("name")) == key[0]
                and norm(b.get("author")) == key[1]
                and norm(b.get("publisher")) == key[2]
            ),
            None,
        )

        if existing_book:
            if not existing_book.get("id"):
                failed.append(
                    {
                        "excel_row": row_num,
                        "name": name,
                        "author": author,
                        "publisher": publisher,
                    }
                )
                tasks_log.append((row_num, "FAILED_UPDATE", name[:50]))
            else:
                tasks_log.append((row_num, "match_db_or_prior", name[:50]))
        else:
            existing.append({"name": name, "author": author, "publisher": publisher})
            tasks_log.append((row_num, "add_stub", name[:50]))

    out = ROOT / "scripts" / "_failed_import_rows.json"
    out.write_text(
        json.dumps({"failed_count": len(failed), "failed": failed}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps({"failed_count": len(failed), "failed": failed}, ensure_ascii=False))
    wb.close()


if __name__ == "__main__":
    main()
